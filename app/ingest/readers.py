from __future__ import annotations

import csv as _csv
import io
import math
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

import pandas as pd

from app.ingest.localizar import MAX_FILAS, Ancla, localizar_cabecera

_BOM_UTF8 = b"\xef\xbb\xbf"
_SEPARADORES = [";", ",", "\t", "|"]

EXTENSIONES_EXCEL = {".xls", ".xlsx", ".xlsm"}
EXTENSIONES_CSV = {".csv", ".txt"}


class ErrorDeLectura(Exception):
    pass


def strip_bom_bytes(data: bytes) -> bytes:
    return data[len(_BOM_UTF8):] if data.startswith(_BOM_UTF8) else data


def decodificar(data: bytes) -> str:
    data = strip_bom_bytes(data)
    try:
        return data.decode("utf-8")
    except UnicodeDecodeError:
        return data.decode("cp1252", errors="replace")


def detectar_separador(primera_linea: str) -> str:
    conteos = {sep: primera_linea.count(sep) for sep in _SEPARADORES}
    mejor = max(conteos, key=conteos.get)
    return mejor if conteos[mejor] > 0 else ";"


def _mejor_separador(lineas: list[str]) -> str:
    mejor, conteo = ";", 0
    for sep in _SEPARADORES:
        total = max((linea.count(sep) for linea in lineas), default=0)
        if total > conteo:
            mejor, conteo = sep, total
    return mejor


def _desenvolver_matriz(matriz: list[list[str]], sep: str) -> list[list[str]]:
    if not matriz or any(len(fila) > 1 for fila in matriz):
        return matriz

    interno = _mejor_separador([fila[0] for fila in matriz if fila])
    if interno == sep and not any(sep in fila[0] for fila in matriz if fila):
        return matriz

    lector = _csv.reader(
        [fila[0] for fila in matriz if fila], delimiter=interno, quotechar='"'
    )
    return [list(fila) for fila in lector]


def _matriz_csv(path: Path, limite: int | None = None) -> list[list[str]]:
    texto = decodificar(path.read_bytes())
    if not texto.strip():
        raise ErrorDeLectura(f"El archivo está vacío: {path.name}")

    lineas = texto.splitlines()
    sep = _mejor_separador(lineas[:MAX_FILAS])

    lector = _csv.reader(io.StringIO(texto), delimiter=sep, quotechar='"')
    filas: list[list[str]] = []
    for fila in lector:
        filas.append([str(c) for c in fila])
        if limite is not None and len(filas) >= limite:
            break

    return _desenvolver_matriz(filas, sep)


def _matriz_xlsx(path: Path, limite: int | None = None) -> list[list[str]]:
    from openpyxl import load_workbook

    libro = load_workbook(path, read_only=True, data_only=True)
    try:
        hoja = _hoja_principal(libro)
        filas: list[list[str]] = []
        for cruda in hoja.iter_rows(max_row=limite, values_only=True):
            filas.append([texto_celda(v) for v in cruda])
            if limite is not None and len(filas) >= limite:
                break
    finally:
        libro.close()

    return filas


def _matriz_xls(path: Path, limite: int | None = None) -> list[list[str]]:
    df = pd.read_excel(
        path, dtype=object, header=None, keep_default_na=False, na_filter=False,
        nrows=limite,
    )
    return [[texto_celda(v) for v in fila] for fila in df.values.tolist()]


def matriz_cruda(path: Path, limite: int | None = None) -> list[list[str]]:
    ext = path.suffix.lower()
    if ext in EXTENSIONES_CSV:
        return _matriz_csv(path, limite)
    if ext == ".xls":
        return _matriz_xls(path, limite)
    if ext in EXTENSIONES_EXCEL:
        return _matriz_xlsx(path, limite)
    raise ErrorDeLectura(f"Formato no soportado: {ext}")


def _ancla(matriz: list[list[str]], esperadas: list[str] | None) -> Ancla:
    if not esperadas:
        return Ancla(0, 0, 0, 0)
    return localizar_cabecera(matriz, esperadas)


def _cabeceras_en(matriz: list[list[str]], ancla: Ancla) -> list[str]:
    if ancla.fila >= len(matriz):
        return []
    fila = matriz[ancla.fila][ancla.columna:]
    return _cabeceras_unicas([texto_celda(c) for c in _recortar_cola_vacia(fila)])


_ENTERO_CON_DECIMAL_CERO = re.compile(r"[+-]?\d+\.0+")
_CIENTIFICA = re.compile(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+")


def normalizar_numero_texto(texto: str) -> str:
    if "." not in texto and "e" not in texto and "E" not in texto:
        return texto

    limpio = texto.strip()
    if not limpio:
        return texto

    if _ENTERO_CON_DECIMAL_CERO.fullmatch(limpio):
        return limpio.split(".", 1)[0]

    if _CIENTIFICA.fullmatch(limpio):
        try:
            numero = Decimal(limpio)
        except Exception:
            return texto
        if numero == numero.to_integral_value():
            return str(int(numero))
        return format(numero, "f")

    return texto


def texto_celda(valor: object) -> str:
    if valor is None:
        return ""

    if isinstance(valor, str):
        return normalizar_numero_texto(valor)

    if isinstance(valor, bool):
        return "TRUE" if valor else "FALSE"

    if isinstance(valor, datetime):
        sin_hora = (
            valor.hour == 0 and valor.minute == 0
            and valor.second == 0 and valor.microsecond == 0
        )
        return valor.strftime("%Y-%m-%d" if sin_hora else "%Y-%m-%d %H:%M:%S")
    if isinstance(valor, date):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, time):
        return valor.strftime("%H:%M:%S")

    if isinstance(valor, int):
        return str(valor)

    if isinstance(valor, float):
        if math.isnan(valor) or math.isinf(valor):
            return ""
        if valor.is_integer() and abs(valor) < 1e16:
            return str(int(valor))
        return _sin_notacion_cientifica(repr(valor))

    if isinstance(valor, Decimal):
        if valor == valor.to_integral_value():
            return str(int(valor))
        return _sin_notacion_cientifica(format(valor.normalize(), "f"))

    return str(valor)


def _sin_notacion_cientifica(texto: str) -> str:
    if "e" not in texto.lower():
        return texto
    try:
        return format(Decimal(texto), "f")
    except Exception:
        return texto


def _cabeceras_unicas(cabeceras: list[str]) -> list[str]:
    vistas: dict[str, int] = {}
    salida: list[str] = []
    for indice, bruta in enumerate(cabeceras):
        nombre = bruta.strip() or f"Columna_{indice + 1}"
        if nombre in vistas:
            vistas[nombre] += 1
            nombre = f"{nombre}.{vistas[nombre]}"
        else:
            vistas[nombre] = 0
        salida.append(nombre)
    return salida


def _recortar_cola_vacia(valores: list) -> list:
    fin = len(valores)
    while fin > 0 and (valores[fin - 1] is None or str(valores[fin - 1]).strip() == ""):
        fin -= 1
    return valores[:fin]


def _hoja_principal(libro):
    hoja = libro[libro.sheetnames[0]]
    if hasattr(hoja, "reset_dimensions"):
        hoja.reset_dimensions()
    return hoja


def _desde_matriz(matriz: list[list[str]], ancla: Ancla) -> pd.DataFrame:
    cabeceras = _cabeceras_en(matriz, ancla)
    if not cabeceras:
        raise ErrorDeLectura("El archivo no tiene cabecera.")

    ancho = len(cabeceras)
    datos: list[list[str]] = []
    for fila in matriz[ancla.fila + 1:]:
        celdas = [texto_celda(v) for v in fila[ancla.columna:]][:ancho]
        if len(celdas) < ancho:
            celdas += [""] * (ancho - len(celdas))
        if any(c.strip() for c in celdas):
            datos.append(celdas)

    return pd.DataFrame(datos, columns=cabeceras, dtype=object)


def leer_como_texto(
    path: str | Path, esperadas: list[str] | None = None
) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        raise ErrorDeLectura(f"No existe el archivo: {path}")

    matriz = matriz_cruda(path)
    if not matriz:
        raise ErrorDeLectura(f"El archivo está vacío: {path.name}")

    df = _desde_matriz(matriz, _ancla(matriz, esperadas))
    df.columns = [str(c) for c in df.columns]
    df = df.map(texto_celda) if hasattr(df, "map") else df.applymap(texto_celda)
    return df


def leer_cabeceras(
    path: str | Path, esperadas: list[str] | None = None
) -> list[str]:
    path = Path(path)
    matriz = matriz_cruda(path, limite=MAX_FILAS)
    if not matriz:
        raise ErrorDeLectura(f"El archivo no tiene cabecera: {path.name}")

    cabeceras = _cabeceras_en(matriz, _ancla(matriz, esperadas))
    if not cabeceras:
        raise ErrorDeLectura(f"El archivo no tiene cabecera: {path.name}")
    return cabeceras
